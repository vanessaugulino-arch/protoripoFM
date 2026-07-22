#!/usr/bin/env python3
"""
Import script — The Fashion Office (contato@thefashionoffice.com.br)
Carrega os dados das 4 abas da planilha para o Supabase.

Uso:
  pip install supabase openpyxl --break-system-packages
  SUPABASE_URL=https://tlbfvuqzvpolfrjwiofx.supabase.co \
  SUPABASE_SERVICE_KEY=<sua_service_role_key> \
  EXCEL_PATH=/caminho/para/"base icloud import.xlsx" \
  python3 import_data_tfo.py

A service_role key está no painel Supabase → Settings → API → service_role.
"""

import os
import sys
import hashlib
from datetime import date, timedelta
from supabase import create_client, Client

# ─── CONFIG ────────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://tlbfvuqzvpolfrjwiofx.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
EXCEL_PATH   = os.environ.get("EXCEL_PATH", "base icloud import.xlsx")
TENANT_ID    = "510da940-e4b4-4750-9b46-fe432bf77065"
BATCH_SIZE   = 200

# Janelas de entrega fictícias por coleção (para data_ultima_entrada)
ENTRADA_WINDOWS = {
    "PV24": ("2023-11-01", "2024-02-28"),
    "OI24": ("2024-04-01", "2024-07-31"),
    "PV25": ("2024-11-01", "2025-02-28"),
    "OI25": ("2025-04-01", "2025-07-31"),
    "PV26": ("2025-11-01", "2026-02-28"),
}

def fictional_entry_date(sku: str, colecao: str) -> str:
    """Gera data fictícia determinística baseada no SKU e coleção."""
    window = ENTRADA_WINDOWS.get(colecao, ("2024-01-01", "2024-06-30"))
    start = date.fromisoformat(window[0])
    end   = date.fromisoformat(window[1])
    days  = max((end - start).days, 1)
    h     = int(hashlib.md5(sku.encode()).hexdigest()[:8], 16)
    return (start + timedelta(days=h % days)).isoformat()

def safe_float(v, default=None):
    try:
        return round(float(v), 4) if v is not None else default
    except (ValueError, TypeError):
        return default

def safe_str(v, default=None):
    return str(v).strip() if v is not None else default

def safe_date(v):
    """Converte datetime ou string para ISO date string."""
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    # DD/MM/YYYY
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        d, m, y = s.split("/")
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return s[:10] if len(s) >= 10 else s

def safe_mes_referencia(v):
    """Converte 'MM/YYYY' para 'YYYY-MM'."""
    if v is None:
        return None
    s = str(v).strip()
    if len(s) == 7 and s[2] == "/":
        m, y = s.split("/")
        return f"{y}-{m.zfill(2)}"
    return s

def insert_batch(supabase: Client, table: str, rows: list, upsert_on: list = None):
    """Insere/upserta um batch de rows em uma tabela."""
    if not rows:
        return 0
    try:
        if upsert_on:
            result = supabase.table(table).upsert(rows, on_conflict=",".join(upsert_on)).execute()
        else:
            result = supabase.table(table).insert(rows, count="exact").execute()
        return len(rows)
    except Exception as e:
        print(f"  ⚠️  Erro no batch de {table}: {e}")
        # Tenta inserir um a um para identificar o problemático
        ok = 0
        for row in rows:
            try:
                if upsert_on:
                    supabase.table(table).upsert([row], on_conflict=",".join(upsert_on)).execute()
                else:
                    supabase.table(table).insert([row]).execute()
                ok += 1
            except Exception as e2:
                print(f"    ❌ Falha na row {list(row.values())[:3]}: {e2}")
        return ok

def run_batches(supabase: Client, table: str, rows: list, upsert_on=None, label=""):
    total = len(rows)
    inserted = 0
    for i in range(0, total, BATCH_SIZE):
        chunk = rows[i:i+BATCH_SIZE]
        n = insert_batch(supabase, table, chunk, upsert_on)
        inserted += n
        print(f"  {label} [{inserted}/{total}] ✓", end="\r")
    print(f"  {label} [{inserted}/{total}] ✓")
    return inserted


def main():
    if not SUPABASE_KEY:
        print("❌ SUPABASE_SERVICE_KEY não definida. Veja as instruções no topo do script.")
        sys.exit(1)

    print(f"Conectando ao Supabase: {SUPABASE_URL}")
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    print(f"\nCarregando Excel: {EXCEL_PATH}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"❌ Arquivo não encontrado: {EXCEL_PATH}")
        sys.exit(1)

    # ── 1. CADASTRO DE PRODUTOS → products ────────────────────────────────────
    print("\n📦 [1/4] Produtos (Cadastro de Produtos → products)...")
    ws = wb["Cadastro de Produtos"]
    rows_excel = list(ws.iter_rows(values_only=True))
    # cols: SKU, Data Última Entrada, Divisão, Categoria, Subcategoria, Linha, Coleção, Preço Tabela, Custo
    products = []
    for r in rows_excel[1:]:
        if not r[0]:
            continue
        sku     = safe_str(r[0])
        divisao = safe_str(r[2])
        categ   = safe_str(r[3])
        subcat  = safe_str(r[4])
        linha   = safe_str(r[5])
        colecao = safe_str(r[6])
        preco   = safe_float(r[7])
        custo   = safe_float(r[8])
        data_e  = fictional_entry_date(sku, colecao or "")
        products.append({
            "tenant_id":           TENANT_ID,
            "sku":                 sku,
            "name":                f"{categ or ''} {subcat or ''} {sku}".strip(),
            "division":            divisao,
            "category":            categ,
            "subcategory":         subcat,
            "linha":               linha,
            "collection_name":     colecao,
            "price_sale":          preco,
            "price_cost":          safe_float(r[8], 0),
            "data_ultima_entrada": data_e,
            "source":              "import",
        })
    run_batches(supabase, "products", products, upsert_on=["sku", "tenant_id"], label="produtos")

    # ── 2. BASE DE PEDIDOS → purchase_orders ──────────────────────────────────
    print("\n🛒 [2/4] Pedidos (Base de Pedidos → purchase_orders)...")
    ws = wb["Base de Pedidos"]
    rows_excel = list(ws.iter_rows(values_only=True))
    # cols: ID_Pedido, Data_Pedido, Data_Entrega_Prevista, Cod_produto, Qtd_Pedida,
    #       Custo_Unitario, Status, Temporada, Mês, Ano
    pedidos = []
    for r in rows_excel[1:]:
        if not r[0]:
            continue
        order_num = safe_str(r[0])
        sku       = safe_str(r[3])
        colecao   = safe_str(r[6])  # Status is col 6, Temporada is col 7
        temporada = safe_str(r[7])
        pedidos.append({
            "tenant_id":        TENANT_ID,
            "order_number":     order_num,
            "sku":              sku,
            "order_date":       safe_date(r[1]) or "2024-01-01",
            "expected_delivery": safe_date(r[2]),
            "quantity_ordered":  int(r[4]) if r[4] else 0,
            "quantity_delivered": 0,
            "unit_cost":        safe_float(r[5]),
            "status":           safe_str(r[6]) or "Pendente",
            "temporada":        temporada,
            "colecao":          temporada,   # In this dataset, colecao = temporada code (PV24, etc.)
            "type":             "purchase",
        })
    run_batches(supabase, "purchase_orders", pedidos, upsert_on=["order_number", "tenant_id"], label="pedidos")

    # ── 3. POSIÇÕES DE ESTOQUE → inventory_snapshots ──────────────────────────
    print("\n📊 [3/4] Estoque (Posições de Estoque → inventory_snapshots)...")
    ws = wb["Posições de Estoque"]
    # cols: Mes_Referencia, Cod_produto, Desc_produto, Departamento, Categoria, Grupo,
    #       Subgrupo, Modelo, Cor, Material, Preco_venda, Custo_produto, Temporada,
    #       Coleção, Qtd_Estoque_Inicial, Mês, Ano, Custo total, Valor de venda total,
    #       Data_ult_entrada, [extras...]
    snapshots = []
    count_total = 0
    for r in ws.iter_rows(values_only=True):
        if count_total == 0:
            count_total += 1
            continue  # skip header
        if not r[1]:  # SKU is col index 1
            continue
        mes_ref   = safe_mes_referencia(r[0])
        sku       = safe_str(r[1])
        qtd       = int(r[14]) if r[14] else 0
        custo_tot = safe_float(r[17], 0)
        venda_tot = safe_float(r[18], 0)
        temporada = safe_str(r[12])
        colecao   = safe_str(r[13])
        data_ult  = safe_date(r[19]) if r[19] else fictional_entry_date(sku, temporada or "")
        # snapshot_date: derive from Mês/Ano (cols 15, 16) or mes_referencia
        mes_num = str(r[15]).zfill(2) if r[15] else "01"
        ano_num = str(r[16])         if r[16] else "2024"
        snap_date = f"{ano_num}-{mes_num}-01"
        snapshots.append({
            "tenant_id":       TENANT_ID,
            "sku":             sku,
            "snapshot_date":   snap_date,
            "quantity":        qtd,
            "value_cost":      custo_tot,
            "value_sale":      venda_tot,
            "temporada":       temporada,
            "colecao":         colecao,
            "mes_referencia":  mes_ref,
            "data_ult_entrada": data_ult,
            "location":        "principal",
        })
        count_total += 1
        if count_total % 5000 == 0:
            print(f"  lendo estoque... {count_total} linhas", end="\r")

    print(f"  {count_total-1} linhas lidas do estoque")
    run_batches(supabase, "inventory_snapshots", snapshots,
                upsert_on=["sku", "tenant_id", "snapshot_date"], label="snapshots")

    # ── 4. VENDAS → sales_history ─────────────────────────────────────────────
    print("\n💰 [4/4] Vendas (VENDAS → sales_history)...")
    ws = wb["VENDAS"]
    # cols (0-based):
    #  0=Data da venda   1=Tipo              2=Código do produto  3=Quantidade
    #  4=Valor Bruto     5=Desconto Aplicado 6=Valor c/ Desconto  7=Imposto
    #  8=Venda Líquida   9=Canal            10=Coleção do Produto 11=Temporada Ativa
    # 12=Forma pagamento 13=Parcelas         14=Mês               15=Ano
    # 16=Divisão        17=Categoria        18=Subcategoria       19=Linha   20=Custo
    sales = []
    count_v = 0
    for r in ws.iter_rows(values_only=True):
        if count_v == 0:
            count_v += 1
            continue  # skip header
        if not r[2]:  # SKU
            continue
        sale_date = safe_date(r[0])
        if not sale_date:
            continue
        rb     = safe_float(r[4], 0)   # Valor Bruto (RB)
        desc   = safe_float(r[5], 0)   # Desconto
        rv     = safe_float(r[6], 0)   # Valor da Venda com desconto (RV)
        imp    = safe_float(r[7], 0)   # Imposto
        rl     = safe_float(r[8], 0)   # Venda Líquida pós imposto (RL)
        qty    = int(r[3]) if r[3] else 0
        colecao   = safe_str(r[10])
        temporada = safe_str(r[11])
        custo_venda = safe_float(r[20])
        sales.append({
            "tenant_id":           TENANT_ID,
            "sku":                 safe_str(r[2]),
            "sale_date":           sale_date,
            "type":                safe_str(r[1]) or "Venda",
            "quantity":            qty,
            "revenue_gross":       rb,
            "discount_value":      desc,
            "revenue_net":         rv,             # RV = pós-desconto, pré-imposto
            "tax_value":           imp,
            "revenue_net_post_tax": rl,            # RL = pós-imposto
            "price_realized":      round(rv / qty, 4) if qty and rv else None,  # PMV
            "channel":             safe_str(r[9]),
            "colecao":             colecao,
            "temporada":           temporada,
            "payment_method":      safe_str(r[12]),
            "installments":        int(r[13]) if r[13] else None,
            "mes":                 safe_str(r[14]),
            "ano":                 int(r[15]) if r[15] else None,
            "category":            safe_str(r[17]),
        })
        count_v += 1
        if count_v % 5000 == 0:
            print(f"  lendo vendas... {count_v} linhas", end="\r")

    print(f"  {count_v-1} linhas lidas de vendas")
    run_batches(supabase, "sales_history", sales, label="vendas")

    wb.close()

    print("\n✅ Import concluído!")
    print(f"   Tenant: {TENANT_ID} (contato@thefashionoffice.com.br)")


if __name__ == "__main__":
    main()

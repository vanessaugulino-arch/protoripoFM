"""
import_catalog_local.py
Importa o cadastro de produtos → Supabase (tabela public.products)

Origem: tabela_estoque_mensal.xlsx  (mesma planilha usada no import de estoque)
        Cada SKU entra uma vez, com os atributos da posição mais recente.

Como usar:
  1. Deixe este script na raiz do projeto, junto com a planilha
  2. No terminal:  python3 import_catalog_local.py
  3. Se interromper, rode de novo — ele faz upsert, não duplica.

Depois de rodar, os cards de Produtos, Hierarquia, Temporadas e Coleções
em Configurações de Operação passam a exibir os dados.

Requisitos: Python 3.8+ e openpyxl  (pip install openpyxl)
"""

import json
import sys
import time
import urllib.error
import urllib.request

# ── Configuração ──────────────────────────────────────────────────────────────
XLSX_PATH = "tabela_estoque_mensal.xlsx"
TENANT_ID = "510da940-e4b4-4750-9b46-fe432bf77065"
SB_URL = "https://tlbfvuqzvpolfrjwiofx.supabase.co"

# Chave anônima — lida do .env do projeto
def read_anon_key() -> str:
    try:
        with open(".env", "r", encoding="utf-8") as fh:
            for line in fh:
                if line.startswith("VITE_SUPABASE_ANON_KEY="):
                    return line.split("=", 1)[1].strip()
    except FileNotFoundError:
        pass
    sys.exit("ERRO: não encontrei VITE_SUPABASE_ANON_KEY no arquivo .env")

SB_KEY = read_anon_key()
BATCH = 300


# ── Leitura da planilha ───────────────────────────────────────────────────────
def load_products():
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERRO: instale a dependência com:  pip install openpyxl")

    print(f"Lendo {XLSX_PATH} ...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    required = ["Cod_produto", "Desc_produto", "Departamento", "Categoria",
                "Grupo", "Subgrupo", "Modelo", "Cor", "Material",
                "Preco_venda", "Custo_produto", "Temporada", "Mes_Referencia"]
    missing = [c for c in required if c not in idx]
    if missing:
        sys.exit(f"ERRO: colunas ausentes na planilha: {missing}")

    def cell(r, name):
        v = r[idx[name]]
        return v

    # Normaliza Mes_Referencia para string ordenável (a planilha mistura
    # datetime e texto dependendo de como a célula foi salva)
    def ref_key(v):
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()[:10]

    # Mantém o registro mais recente de cada SKU
    latest = {}
    total = 0
    for r in rows:
        sku = cell(r, "Cod_produto")
        if not sku:
            continue
        total += 1
        ref = ref_key(cell(r, "Mes_Referencia"))
        prev = latest.get(sku)
        if prev is None or ref >= prev[0]:
            latest[sku] = (ref, r)

    print(f"  {total} linhas lidas · {len(latest)} SKUs únicos")

    def txt(v):
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    def num(v):
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return None

    out = []
    for sku, (_ref, r) in latest.items():
        temporada = txt(cell(r, "Temporada"))
        out.append({
            "tenant_id": TENANT_ID,
            "sku": str(sku).strip(),
            "name": txt(cell(r, "Desc_produto")) or str(sku).strip(),
            "division": txt(cell(r, "Departamento")),
            "category": txt(cell(r, "Categoria")),
            "subcategory": txt(cell(r, "Subgrupo")),
            "linha": txt(cell(r, "Grupo")),
            "model": txt(cell(r, "Modelo")),
            "color": txt(cell(r, "Cor")),
            "material": txt(cell(r, "Material")),
            "price_sale": num(cell(r, "Preco_venda")),
            "price_cost": num(cell(r, "Custo_produto")),
            "season": temporada,
            "collection_name": temporada,
            "source": "planilha",  # constraint products_source_check: manual|planilha|erp
            "attributes": {},
        })
    return out


# ── Envio ao Supabase ─────────────────────────────────────────────────────────
def push(records):
    url = f"{SB_URL}/rest/v1/products?on_conflict=tenant_id,sku"
    headers = {
        "apikey": SB_KEY,
        "Authorization": f"Bearer {SB_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    sent = 0
    failed = 0
    for i in range(0, len(records), BATCH):
        chunk = records[i:i + BATCH]
        body = json.dumps(chunk, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers=headers, method="POST")
        for attempt in range(3):
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    if resp.status < 300:
                        sent += len(chunk)
                        break
            except urllib.error.HTTPError as e:
                detail = e.read().decode("utf-8", "replace")[:300]
                if attempt == 2:
                    failed += len(chunk)
                    print(f"  ! lote {i//BATCH + 1} falhou: {e.code} {detail}")
                else:
                    time.sleep(1.5 * (attempt + 1))
            except Exception as e:  # rede instável
                if attempt == 2:
                    failed += len(chunk)
                    print(f"  ! lote {i//BATCH + 1} falhou: {e}")
                else:
                    time.sleep(1.5 * (attempt + 1))
        done = min(i + BATCH, len(records))
        print(f"  {done}/{len(records)} enviados", end="\r")
    print()
    return sent, failed


def push_hierarquia(records):
    """
    Deriva divisão → categoria → subcategoria dos produtos e grava em
    public.hierarquia_produtos, que alimenta o card de Hierarquia de Produtos.
    """
    seen = []
    key = set()
    for r in records:
        div, cat, sub = r["division"], r["category"], r["subcategory"]
        if not div or not cat:
            continue
        k = (div, cat, sub or "")
        if k in key:
            continue
        key.add(k)
        seen.append({
            "tenant_id": TENANT_ID,
            "divisao": div,
            "categoria": cat,
            "subcategoria": sub or "",
            "ordem": len(seen),
            "ativo": True,
        })

    if not seen:
        return 0, 0

    print(f"Gravando {len(seen)} nós de hierarquia ...")
    headers = {
        "apikey": SB_KEY,
        "Authorization": f"Bearer {SB_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    url = (f"{SB_URL}/rest/v1/hierarquia_produtos"
           "?on_conflict=tenant_id,divisao,categoria,subcategoria")
    body = json.dumps(seen, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            if resp.status < 300:
                return len(seen), 0
    except urllib.error.HTTPError as e:
        print(f"  ! hierarquia falhou: {e.code} {e.read().decode('utf-8','replace')[:200]}")
    except Exception as e:
        print(f"  ! hierarquia falhou: {e}")
    return 0, len(seen)


def main():
    records = load_products()
    if not records:
        sys.exit("Nada a importar.")
    print(f"Enviando {len(records)} produtos para o Supabase ...")
    sent, failed = push(records)
    print(f"Produtos: {sent} gravados · {failed} com falha")

    if sent > 0:
        hs, hf = push_hierarquia(records)
        print(f"Hierarquia: {hs} gravados · {hf} com falha")

    print()
    if failed == 0:
        print("Concluído. Abra Configurações de Operação no sistema — os cards de")
        print("Produtos, Hierarquia, Temporadas e Coleções devem estar preenchidos.")
    else:
        print("Concluído com falhas. Rode novamente para reprocessar os lotes que falharam.")


if __name__ == "__main__":
    main()
